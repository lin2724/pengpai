# coding=utf-8
import os
import sys
import re
import requests
from lxml import etree
from lxml import html
from io import StringIO
from openpyxl import Workbook
from openpyxl import load_workbook
from sqlite_util import DBHandler
from sqlite_util import DBRowHuaBan
from common_lib import MyArgParse
from common_lib import LogHandle

gstLogHandler = LogHandle('pengpai.log')


def clean_str(str_line):
    pattern = '\s*(?P<content>.*)'
    m = re.search(pattern, str_line)
    if m:
        return m.group('content')
    return ''


def match_article_content(content):
    pattern = '^\s(?P<head>\<[A-z]+ ).*'
    m = re.match(pattern=pattern, string=content)
    if m:
        head = m.group('head')
        full_pattern = pattern = '^\s(?P<head>\<%s+ ).*?\</%s\>(?P<rest_content>.*)' % (head, head)
        m = re.match(full_pattern, string=content)
        if not m:
            print 'Wrong pattern'
            return False, None, False
        rest_content = m.group('rest_content')
        return True, rest_content, False
    else:
        pattern = '^(?<![<]>)'


def get_artical_length_from_html(html_content):
    content_length = 0
    article_content = ''
    pattern_list = list()
    pattern = "<div\s+class=\"news_txt\"\s+data-size=\"standard\">(?P<content>[\s\S]*?<br />)"
    pattern_list.append(pattern)
    if '<div class="contheight"></div>' in html_content:
        pattern = "<div\s+class=\"news_txt\"\s+data-size=\"standard\">(?P<content>[\s\S]*?<br />)"
        pattern_list.append(pattern)
    else:
        pattern = "\<div class=\'contheight\'\>\</div\>(?P<content>[\s\S]*?)<div\s+class=\"news_editor\">"
        pattern_list.append(pattern)
    for pattern in pattern_list:
        m = re.findall(pattern=pattern, string=html_content)
        if m:
            print 'find [%d]' % len(m)
            for line in m:
                for s in line.decode('utf-8'):
                    if s > '~' or s < ' ':
                        content_length += 1
                        article_content += s
    return content_length, article_content
    pass


def get_info_from_content(article_content, article_url=''):
    title = ''
    auth = ''
    key_word = ''
    from_src = ''
    article_time = ''
    comments_count = 0
    thumb_up_count = 0

    root = etree.HTML(article_content)
    nodes = root.xpath('//h1[@class="news_title"]')
    if len(nodes):
        title = nodes[0].text
        print title
    else:
        print 'about not found'

    nodes = root.xpath('//div[@class="news_about"]')
    if len(nodes):
        auth = nodes[0].xpath('.//p')[0].text
        info_node = nodes[0].xpath('.//p')[1]
        article_time = clean_str(info_node.text)
        from_src = info_node.xpath('.//span')[0].text
        print auth
        print article_time
        print from_src
    else:
        print 'about not found'

    nodes = root.xpath('//div[@class="news_keyword"]')
    if len(nodes):
        key_word = nodes[0].text
    else:
        print 'keyword not found'

    nodes = root.xpath('//a[@id="zan"]')
    if len(nodes):
        thumb_up_count = clean_str(nodes[0].text)
    else:
        print 'zan not found'

    nodes = root.xpath('//h2[@id="comm_span"]')
    if len(nodes):
        print nodes[0].text
        comments_count = clean_str(nodes[0].xpath('.//span')[0].text)
        print comments_count
    else:
        print 'command not found'

    character_cnt, article_content = get_artical_length_from_html(article_content)

    ret_dict = dict()

    ret_dict['title'] = title
    ret_dict['auth'] = auth
    ret_dict['key_word'] = key_word
    ret_dict['from_src'] = from_src
    ret_dict['article_time'] = article_time
    ret_dict['comments_count'] = comments_count
    ret_dict['thumb_up_count'] = thumb_up_count
    ret_dict['url'] = article_url
    ret_dict['character_cnt'] = character_cnt
    ret_dict['article_content'] = article_content
    print ret_dict
    return ret_dict

    pass


class SaverXLSX:
    def __init__(self, file_path=''):
        if not file_path.endswith('.xlsx') or not file_path.endswith('.xls'):
            file_path += '.xls'
        self.file_path = file_path
        if os.path.exists(file_path):
            self.wb = load_workbook(filename=file_path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.do_init_xlsx()

    def do_init_xlsx(self):
        self.ws.append(['title', 'auth', 'key_word', 'from_src', 'article_time', 'comments_count', 'thumb_up_count',
                        'character_cnt'])
        pass

    def quit(self):
        self.wb.save(filename=self.file_path)

    def insert_row(self, row_dict):
        self.ws.append([row_dict['title'],
                       row_dict['auth'],
                       row_dict['key_word'],
                       row_dict['from_src'],
                       row_dict['article_time'],
                       row_dict['comments_count'],
                       row_dict['thumb_up_count'],
                        row_dict['character_cnt']])
        pass


class SaverDB:
    def __init__(self):
        self.handler = DBHandler()
        self.handler.load('PengPai.db')
        self.handler.add_table('PengPai')
        pass

    def insert_info(self, row_dict):
        huaban_row = DBRowHuaBan()
        huaban_row.item_list[0].value = row_dict['title']
        huaban_row.item_list[1].value = row_dict['auth']
        huaban_row.item_list[2].value = row_dict['key_word']
        huaban_row.item_list[3].value = row_dict['from_src']
        huaban_row.item_list[4].value = row_dict['article_time']
        huaban_row.item_list[5].value = 0 # int(row_dict['comments_count'])
        huaban_row.item_list[6].value = 0 # int(row_dict['thumb_up_count'])
        huaban_row.item_list[7].value = row_dict['url']
        huaban_row.item_list[8].value = row_dict['character_cnt']
        self.handler.insert_row(huaban_row)
        pass


class SaverContentText:
    def __init__(self, save_file_path='article_content.txt'):
        self.log = gstLogHandler.log
        self.m_set_save_file_path = save_file_path
        self.m_set_fd = open(self.m_set_save_file_path, 'w+')
        self.log('save all article txt to [%s]' % self.m_set_save_file_path)
        pass

    def insert_info(self, row_dict):
        info_line = '\n\nurl:[%s]\ntitle[%s] auth[%s] keyword[%s] from[%s] time[%s]\n' \
                    % (row_dict['url'],row_dict['title'], row_dict['auth'], row_dict['key_word'], row_dict['from_src'], row_dict['article_time'])
        self.m_set_fd.write(info_line.encode('utf-8'))
        self.m_set_fd.write(row_dict['article_content'].encode('utf-8'))
        pass

    def quit(self):
        self.m_set_fd.close()


def do_test(file_path):
    with open(file_path, 'r') as fd:
        content = fd.read()
        get_info_from_content(content)
    pass


def do_test_save(file_name):
    saver = SaverXLSX(file_name)
    saver.quit()

    pass


def get_all_file_list(folder_path):
    ret_list = list()
    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            file_path = os.path.join(root, file_name)
            if file_path[-len('.html'):] == '.html':
                ret_list.append(file_path)
                pass
    return ret_list


def get_url_from_html_path(html_path):
    file_name = os.path.basename(html_path)
    pattern = '(?P<article_id>\d+).*'
    m = re.search(pattern=pattern, string=file_name)
    if m:
        url = 'https://www.thepaper.cn/newsDetail_forward_%s' % m.group('article_id')
        return url
    else:
        print 'Failed get url'
        pass
    return ''
    pass


def do_parser_info(folder_path, xlsx_name):
    saver = SaverXLSX(xlsx_name)
    DBSaver = SaverDB()
    ContentSaver = SaverContentText()
    html_list = get_all_file_list(folder_path)
    for html_file in html_list:
        with open(html_file, 'r') as fd:
            content = fd.read()
            row_dict = get_info_from_content(content, get_url_from_html_path(html_file))
            saver.insert_row(row_dict)
            DBSaver.insert_info(row_dict)
            ContentSaver.insert_info(row_dict)
    saver.quit()


def init_argparser():
    arg_parse = MyArgParse()
    arg_parse.add_option('-savecontent', 1, 'save all content to a file')
    arg_parse.add_option('-d', 1, 'folder where articles are stored')
    arg_parse.add_option('-h', 0, 'print help')
    return arg_parse
    pass

if __name__ == '__main__':
    if len(sys.argv) == 2:
        print 'Usage:python %s article_folder xls_file_name'
        get_all_file_list(sys.argv[1])
        exit(0)
        do_test_save(sys.argv[1])
        exit(0)
        do_test(sys.argv[1])
    elif len(sys.argv) == 3:
        do_parser_info(sys.argv[1], sys.argv[2])











